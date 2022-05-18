from analyze.schema import *
from typing import List
import openpyxl

init_params = InitialParameters()
main_fittings: List[MainFitting] = list()
pump2_fittings: List[Pump2Fitting] = list()
pump3_fittings: List[Pump3Fitting] = list()
wet_bulb_fittings_1to1: List[WetBulbFitting_1to1] = list()
wet_bulb_fittings_2to1: List[WetBulbFitting_2to1] = list()
wet_bulb_fittings_3to1: List[WetBulbFitting_3to1] = list()
wet_bulb_fittings_4to1: List[WetBulbFitting_4to1] = list()
wet_bulb_fittings_3to2: List[WetBulbFitting_3to2] = list()
wet_bulb_fittings_4to3: List[WetBulbFitting_4to3] = list()
p4_fittings: List[P4Fitting] = list()

fitting_coefficients = FittingCoefficients()  # 拟合结果
optimize_result: List[OptimizeResult] = list()  # 优化结果
q_delta: List[QDeltaEntry] = list()  # Q值变化


def load(path: str):
    global init_params, main_fittings, pump2_fittings, pump3_fittings, wet_bulb_fittings_1to1, wet_bulb_fittings_2to1, wet_bulb_fittings_3to1, wet_bulb_fittings_4to1, wet_bulb_fittings_3to2, wet_bulb_fittings_4to3, p4_fittings, \
        fitting_coefficients, optimize_result, q_delta
    wb = openpyxl.load_workbook(path, read_only=True)
    for sheet in wb:
        if sheet.title == "q值变化值":
            q_delta = list()
            optimize_result = list()
            for idx, row in enumerate(sheet.rows):
                if idx <= 0:
                    continue
                entry = QDeltaEntry()
                entry.year = row[0].value
                entry.mon = row[1].value
                entry.day = row[2].value
                entry.hour = row[3].value
                entry.Q = row[4].value
                entry.Ts = row[5].value
                entry.T = row[6].value
                q_delta.append(entry)
                temp = OptimizeResult()
                temp.year = row[0].value
                temp.mon = row[1].value
                temp.day = row[2].value
                temp.hour = row[3].value
                temp.q = float(row[4].value)
                temp.ts = float(row[5].value)
                optimize_result.append(temp)
                # print(len(optimize_result))
        elif sheet.title == "参数初始值设定":
            init_params.q = float(sheet["B4"].value)
            init_params.n = int(sheet["B5"].value)
            init_params.efficiency_range = float(sheet["B6"].value)
            init_params.t3_min = float(sheet["B7"].value)
            init_params.G20 = float(sheet["B9"].value)
            init_params.G30 = float(sheet["B14"].value)
            init_params.h = float(sheet["B10"].value)
            init_params.p2 = float(sheet["B11"].value)
            init_params.delta_t1_range = (float(sheet["B24"].value), float(sheet["C24"].value))
            init_params.delta_t2_range = (float(sheet["B25"].value), float(sheet["C25"].value))
            init_params.q_min = float(sheet["B26"].value)
            # init_params.p20 = float(sheet["B27"].value)
            init_params.P0 = float(sheet["B21"].value)
            init_params.mu = float(sheet["B27"].value)
            init_params.lamb = float(sheet["B28"].value)
            init_params.lengque_maxn = float(sheet["B22"].value)
            init_params.yuzhi = float(sheet["B29"].value)

            for i in range(9):
                init_params.t1_range.append(float(sheet.cell(2, i + 2).value))

        elif sheet.title == "主机参数拟合":
            main_fittings = list()
            for rowidx, row in enumerate(sheet.rows):
                if rowidx <= 1:
                    continue
                if row[1].value is None:
                    break
                entry = MainFitting()
                entry.q = float(row[1].value)
                entry.p1 = float(row[2].value)
                entry.t1 = float(row[3].value)
                entry.t2 = float(row[4].value)
                entry.t3 = float(row[5].value)
                entry.t4 = float(row[6].value)
                entry.cop = float(row[7].value)
                main_fittings.append(entry)
        elif sheet.title == "水泵性能参数拟合":
            pump2_fittings = list()
            for i in range(5):
                entry = Pump2Fitting()
                entry.g2 = float(sheet.cell(4, 3 + i * 2).value)
                entry.p2 = float(sheet.cell(4, 4 + i * 2).value)
                pump2_fittings.append(entry)

            pump3_fittings = list()
            for i in range(5):
                entry = Pump3Fitting()
                entry.g3 = float(sheet.cell(11, 3 + i * 2).value)
                entry.p3 = float(sheet.cell(11, 4 + i * 2).value)
                pump3_fittings.append(entry)
        elif sheet.title == "冷却塔拟合":
            i = 0
            wet_bulb_fittings_1to1 = list()
            while sheet.cell(3 + i, 1).value is not None:
                entry = WetBulbFitting_1to1()
                entry.temp = float(sheet.cell(3 + i, 1).value)
                entry.amplitude = float(sheet.cell(3 + i, 2).value)
                wet_bulb_fittings_1to1.append(entry)
                i += 1

            i = 0
            wet_bulb_fittings_2to1 = list()
            while sheet.cell(3 + i, 3).value is not None:
                entry = WetBulbFitting_2to1()
                entry.temp = float(sheet.cell(3 + i, 3).value)
                entry.amplitude = float(sheet.cell(3 + i, 4).value)
                wet_bulb_fittings_2to1.append(entry)
                i += 1

            i = 0
            wet_bulb_fittings_3to1 = list()
            while sheet.cell(3 + i, 5).value is not None:
                entry = WetBulbFitting_3to1()
                entry.temp = float(sheet.cell(3 + i, 5).value)
                entry.amplitude = float(sheet.cell(3 + i, 6).value)
                wet_bulb_fittings_3to1.append(entry)
                i += 1

            i = 0
            wet_bulb_fittings_4to1 = list()
            while sheet.cell(3 + i, 7).value is not None:
                entry = WetBulbFitting_4to1()
                entry.temp = float(sheet.cell(3 + i, 7).value)
                entry.amplitude = float(sheet.cell(3 + i, 8).value)
                wet_bulb_fittings_4to1.append(entry)
                i += 1

            i = 0
            wet_bulb_fittings_3to2 = list()
            while sheet.cell(3 + i, 9).value is not None:
                entry = WetBulbFitting_3to2()
                entry.temp = float(sheet.cell(3 + i, 9).value)
                entry.amplitude = float(sheet.cell(3 + i, 10).value)
                wet_bulb_fittings_3to2.append(entry)
                i += 1

            i = 0
            wet_bulb_fittings_4to3 = list()
            while sheet.cell(3 + i, 11).value is not None:
                entry = WetBulbFitting_4to3()
                entry.temp = float(sheet.cell(3 + i, 11).value)
                entry.amplitude = float(sheet.cell(3 + i, 12).value)
                wet_bulb_fittings_4to3.append(entry)
                i += 1

            i = 0
            p4_fittings = list()
            while sheet.cell(3 + i, 13).value is not None:
                entry = P4Fitting()
                entry.g = float(sheet.cell(3 + i, 13).value)
                entry.p4 = float(sheet.cell(3 + i, 14).value)
                p4_fittings.append(entry)
                i += 1

        elif sheet.title == "拟合系数表":
            fitting_coefficients = FittingCoefficients()
            for i in range(12):
                val = 0.0
                if sheet.cell(3, 2 + i).value is not None:
                    val = float(sheet.cell(3, 2 + i).value)
                fitting_coefficients.b.append(val)
            for i in range(12):
                val = 0.0
                if sheet.cell(5, 2 + i).value is not None:
                    val = float(sheet.cell(3, 2 + i).value)
                fitting_coefficients.b.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(8, 2 + i).value is not None:
                    val = float(sheet.cell(8, 2 + i).value)
                fitting_coefficients.a.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(11, 2 + i).value is not None:
                    val = float(sheet.cell(11, 2 + i).value)
                fitting_coefficients.c.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(15, 2 + i).value is not None:
                    val = float(sheet.cell(15, 2 + i).value)
                fitting_coefficients.d_1to1.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(17, 2 + i).value is not None:
                    val = float(sheet.cell(17, 2 + i).value)
                fitting_coefficients.d_2to1.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(19, 2 + i).value is not None:
                    val = float(sheet.cell(19, 2 + i).value)
                fitting_coefficients.d_3to1.append(val)
            for i in range(4):
                val = 0.0
                if sheet.cell(21, 2 + i).value is not None:
                    val = float(sheet.cell(21, 2 + i).value)
                fitting_coefficients.d_4to1.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(23, 2 + i).value is not None:
                    val = float(sheet.cell(23, 2 + i).value)
                fitting_coefficients.d_3to2.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(25, 2 + i).value is not None:
                    val = float(sheet.cell(25, 2 + i).value)
                fitting_coefficients.d_4to3.append(val)
            for i in range(4):
                val = 0.0
                if sheet.cell(27, 2 + i).value is not None:
                    val = float(sheet.cell(27, 2 + i).value)
                fitting_coefficients.e.append(val)
    print("load completed!")
    wb.close()
